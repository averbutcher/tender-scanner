import re
import json
from datetime import datetime, timedelta, time as time_type
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CONFIG_PATH = Path(__file__).parent / "config.json"

DEFAULT_CONFIG = {
    "excel_columns": {
        "date": "B",
        "start_time": "C",
        "end_time": "D",
        "worker_name": "F"
    },
    "excel_has_header": True,
    "rules": {
        "gap_threshold_minutes": 30,
        "default_start_time": "10:00"
    },
    "aliases": {}
}


def load_config():
    if not CONFIG_PATH.exists():
        save_config(DEFAULT_CONFIG)
        return dict(DEFAULT_CONFIG)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def col_to_idx(letter: str) -> int:
    return ord(letter.strip().upper()) - ord("A")


def _is_na(val) -> bool:
    try:
        return bool(pd.isna(val))
    except (TypeError, ValueError):
        return False


def _date_matches(excel_date, day: int, month: int) -> bool:
    """Return True if excel_date falls on the given day and month."""
    if excel_date is None or _is_na(excel_date):
        return False
    try:
        return excel_date.day == day and excel_date.month == month
    except AttributeError:
        pass
    if isinstance(excel_date, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d.%m.%Y"):
            try:
                dt = datetime.strptime(excel_date.strip(), fmt)
                return dt.day == day and dt.month == month
            except ValueError:
                pass
    return False


def parse_time(val) -> time_type | None:
    if val is None or _is_na(val):
        return None
    if isinstance(val, time_type):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(val.strip(), fmt).time()
            except ValueError:
                pass
        return None
    if isinstance(val, (int, float)):
        # Excel stores times as fractions of a day
        if 0 <= val <= 1:
            total_sec = round(val * 86400)
            h, rem = divmod(total_sec, 3600)
            m = rem // 60
            return time_type(h % 24, m)
    return None


def time_to_hours(t: time_type) -> float:
    return t.hour + t.minute / 60 + t.second / 3600


def hours_between(start: time_type, end: time_type) -> float:
    diff = time_to_hours(end) - time_to_hours(start)
    if diff < 0:
        diff += 24  # overnight shift
    return diff


def add_hours(t: time_type, h: float) -> time_type:
    dt = datetime.combine(datetime.today(), t) + timedelta(hours=h)
    return dt.time()


def subtract_hours(t: time_type, h: float) -> time_type:
    dt = datetime.combine(datetime.today(), t) - timedelta(hours=h)
    return dt.time()


# ── Pre-validation ────────────────────────────────────────────────────────────

def find_suspicious_lines(text: str) -> list[str]:
    """
    Return lines that look like worker entries but are missing commas.
    A valid worker line needs 2 commas (name, workplace, hours).
    We flag lines with <2 commas that contain Hebrew text AND either
    at least 1 comma or an hours/sales indicator (ש / מ + digit).
    Pure date headers are excluded.
    """
    suspicious = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        # Skip pure date headers like "סופי 15.6" or "ראשון 22/6"
        if re.search(r'\d{1,2}[./]\d{1,2}', line) and line.count(',') == 0:
            continue
        if line.count(',') >= 2:
            continue  # Valid format, not suspicious
        has_hebrew = bool(re.search(r'[א-ת]', line))
        has_hours  = bool(re.search(r'\d+ש|ש\d+|\d+מ|מ\d+', line))
        if has_hebrew and (line.count(',') >= 1 or has_hours):
            suspicious.append(line)
    return suspicious


# ── Message parsing ────────────────────────────────────────────────────────────

def parse_message(text: str) -> list[dict]:
    entries     = []
    current_date = None  # (day, month) tuple, set when a date header is found

    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue

        parts = line.split(",", 2)

        if len(parts) < 3:
            # Not a worker line — check for a date header (any format: DD.M / DD/M)
            # Negative lookahead (?!\d*[שמ]) prevents matching decimal hours like 8.5ש
            date_m = re.search(r'(\d{1,2})[./](\d{1,2})(?!\d*[שמ])', line)
            if date_m:
                current_date = (int(date_m.group(1)), int(date_m.group(2)))
            continue

        worker_name = parts[0].strip()
        workplace   = parts[1].strip()
        rest        = parts[2].strip()

        # Detect direction: number-then-letter (8ש) or letter-then-number (ש8)
        if re.search(r"\d+ש", rest):
            hours_m = re.search(r"(\d+(?:\.\d+)?)ש", rest)
            sales_m = re.search(r"(\d+(?:\.\d+)?)מ", rest)
            hours = float(hours_m.group(1)) if hours_m else None
            sales = int(float(sales_m.group(1))) if sales_m else ""
        else:
            hours_m = re.search(r"ש(\d+(?:\.\d+)?)", rest)
            sales_m = re.search(r"מ(\d+(?:\.\d+)?)", rest)
            hours = float(hours_m.group(1)) if hours_m else None
            sales = int(float(sales_m.group(1))) if sales_m else ""

        if worker_name:
            entries.append({
                "worker_name": worker_name,
                "workplace":   workplace,
                "hours":       hours,
                "sales":       sales,
                "msg_date":    current_date,
            })
    return entries


# ── Excel parsing ──────────────────────────────────────────────────────────────

def _is_ignored(name: str, ignored_names: list) -> bool:
    """
    Returns True if 'name' matches any entry in ignored_names.
    Handles word reordering and partial names (e.g. first name only).
    """
    name_words = set(name.strip().split())
    for ignored in ignored_names:
        ig_words = set(ignored.split())
        if not ig_words:
            continue
        # All words of the shorter side appear in the longer side
        if ig_words.issubset(name_words) or name_words.issubset(ig_words):
            return True
    return False


def parse_excel(file_path: str, cfg: dict) -> list[dict]:
    cols          = cfg["excel_columns"]
    has_header    = cfg.get("excel_has_header", True)
    ignored_names = [n.strip() for n in cfg.get("ignored_names", []) if n.strip()]

    date_idx  = col_to_idx(cols["date"])
    start_idx = col_to_idx(cols["start_time"])
    end_idx   = col_to_idx(cols["end_time"])
    name_idx  = col_to_idx(cols["worker_name"])

    skip = 1 if has_header else 0
    df   = pd.read_excel(file_path, header=None, skiprows=skip)

    entries = []
    for _, row in df.iterrows():
        try:
            name = str(row.iloc[name_idx]).strip()
            if not name or name.lower() == "nan":
                continue
            if _is_ignored(name, ignored_names):
                continue

            date_val = row.iloc[date_idx]
            start_t  = parse_time(row.iloc[start_idx])
            end_t    = parse_time(row.iloc[end_idx])

            # Skip rows with no start time (worker forgot to clock in, end time may be misplaced)
            if start_t is None:
                continue

            hours = hours_between(start_t, end_t) if (start_t is not None and end_t is not None) else None

            entries.append({
                "worker_name": name,
                "date":        date_val,
                "start_time":  start_t,
                "end_time":    end_t,
                "hours":       hours,
            })
        except Exception:
            continue
    return entries


# ── Name matching ──────────────────────────────────────────────────────────────

def canonical_key(name: str, aliases: dict) -> str:
    """Map a name to its canonical group key using aliases."""
    for k, v in aliases.items():
        if name == k or name == v:
            return k
    return name


def _first_name(name: str) -> str:
    return name.split()[0] if name.strip() else name


def _build_word_lookup(excel_by_key: dict, word_index: int) -> dict[str, str | None]:
    """
    Build {word: canonical_key} where 'word' is the word at word_index in the
    Excel name. Keys that map to more than one worker are set to None (ambiguous).
    """
    lookup: dict[str, str | None] = {}
    for key in excel_by_key:
        words = key.split()
        if len(words) <= word_index:
            continue
        word = words[word_index]
        if word in lookup:
            lookup[word] = None  # ambiguous
        else:
            lookup[word] = key
    return lookup


# ── Comparison ─────────────────────────────────────────────────────────────────

def _filter_excel_rows(rows: list) -> list:
    """
    For a worker with multiple Excel rows, keep only 'good' rows
    (both times present, shift >= 30 min) if any exist.
    Falls back to all rows when none qualify.
    """
    if len(rows) <= 1:
        return rows
    good = [
        r for r in rows
        if r["start_time"] is not None
        and r["end_time"] is not None
        and r.get("hours") is not None
        and r["hours"] >= 0.5
    ]
    return good if good else rows

def compare(msg_entries: list, excel_entries: list, cfg: dict) -> list[dict]:
    """
    Entry point. Groups message entries by date, runs per-day comparison,
    and concatenates results. Raises ValueError if no date headers are found.
    """
    dated = [e for e in msg_entries if e.get("msg_date") is not None]
    if not dated:
        raise ValueError(
            "לא נמצאה כותרת תאריך בהודעה\n"
            "יש להוסיף שורת תאריך לפני כל יום (לדוגמה: סופי 15.6)"
        )

    # Collect dates in order of first appearance
    seen_dates: list = []
    msg_by_date: dict = {}
    for e in msg_entries:
        d = e["msg_date"]
        if d not in seen_dates:
            seen_dates.append(d)
        msg_by_date.setdefault(d, []).append(e)

    results = []
    for date in seen_dates:
        day_xl = [r for r in excel_entries if _date_matches(r["date"], date[0], date[1])]
        results.extend(_compare_day(msg_by_date[date], day_xl, cfg))
    return results


def _compare_day(msg_entries: list, excel_entries: list, cfg: dict) -> list[dict]:
    aliases       = cfg.get("aliases", {})
    threshold_min = cfg["rules"]["gap_threshold_minutes"]
    default_start = datetime.strptime(cfg["rules"]["default_start_time"], "%H:%M").time()
    managers      = [m.strip() for m in cfg.get("managers", []) if m.strip()]

    # Infer "the date" for this run from Excel
    excel_date = next(
        (e["date"] for e in excel_entries if not _is_na(e["date"])),
        None
    )

    excel_by_key: dict[str, list] = {}
    for e in excel_entries:
        k = canonical_key(e["worker_name"], aliases)
        excel_by_key.setdefault(k, []).append(e)

    # Name fallback lookups (word 0 = first word, word 1 = second word of Excel name)
    first_word_lookup  = _build_word_lookup(excel_by_key, 0)  # Excel first name first
    second_word_lookup = _build_word_lookup(excel_by_key, 1)  # Excel last name first

    def resolve_msg_key(msg_name: str) -> str:
        """
        Matching priority:
        1. Alias
        2. Exact name
        3. Message first word == Excel first word  (e.g. "ישראל" → "ישראל ישראלי")
        4. Message first word == Excel second word (e.g. "ישראל" → "ישראלי ישראל")
        """
        k = canonical_key(msg_name, aliases)
        if k in excel_by_key:
            return k
        fn = _first_name(msg_name)
        matched = first_word_lookup.get(fn)
        if matched is not None:
            return matched
        matched = second_word_lookup.get(fn)
        if matched is not None:
            return matched
        return k

    msg_by_key: dict[str, list] = {}
    for e in msg_entries:
        k = resolve_msg_key(e["worker_name"])
        msg_by_key.setdefault(k, []).append(e)

    all_keys = sorted(set(excel_by_key) | set(msg_by_key))
    output   = []

    for key in all_keys:
        ex_rows = _filter_excel_rows(excel_by_key.get(key, []))
        ms_rows = msg_by_key.get(key, [])

        # ── 1 message row, multiple Excel rows → sum Excel hours ────────────
        if len(ms_rows) == 1 and len(ex_rows) > 1:
            mr = ms_rows[0]
            mh = mr["hours"] or 0
            total_xl_hours = sum(r["hours"] or 0 for r in ex_rows if r["hours"] is not None)
            # Use earliest start and latest end from the Excel rows
            starts = [r["start_time"] for r in ex_rows if r["start_time"] is not None]
            ends   = [r["end_time"]   for r in ex_rows if r["end_time"]   is not None]
            start_t = min(starts) if starts else None
            end_t   = max(ends)   if ends   else None
            # Representative Excel row for date/name
            er = ex_rows[0]
            diff_min = abs(total_xl_hours - mh) * 60
            if diff_min <= threshold_min:
                output.append({
                    "date":        er["date"],
                    "worker_name": er["worker_name"],
                    "workplace":   mr["workplace"],
                    "start_time":  start_t,
                    "end_time":    end_t,
                    "sales":       mr["sales"],
                    "notes":       "הכל תקין",
                    "status":      "ok",
                })
            else:
                output.append({
                    "date":        er["date"],
                    "worker_name": er["worker_name"],
                    "workplace":   mr["workplace"],
                    "start_time":  start_t,
                    "end_time":    end_t,
                    "sales":       mr["sales"],
                    "notes":       f"פער של {int(diff_min)} דקות",
                    "status":      "gap",
                })
            continue
        # ────────────────────────────────────────────────────────────────────

        # ── Multi-workplace: several message rows but one Excel row ──────────
        if len(ms_rows) > 1 and len(ex_rows) == 1:
            er               = ex_rows[0]
            eh               = er["hours"]
            start_t          = er["start_time"]
            end_t            = er["end_time"]
            partial_note     = None
            total_msg_hours  = sum(mr["hours"] or 0 for mr in ms_rows)

            # Complete missing Excel time using total message hours
            if start_t is None and end_t is not None:
                partial_note = "שעת התחלה חסרה באקסל"
                start_t = subtract_hours(end_t, total_msg_hours)
                eh = total_msg_hours
            elif end_t is None and start_t is not None:
                partial_note = "שעת סיום חסרה באקסל"
                eh = total_msg_hours

            diff_min = abs((eh or total_msg_hours) - total_msg_hours) * 60

            if diff_min <= threshold_min:
                status    = "ok"
                base_note = partial_note or "הכל תקין"
            else:
                status    = "gap"
                base_note = partial_note or f"פער של {int(diff_min)} דקות"

            current_start = start_t
            for mr in ms_rows:
                msg_hours = mr["hours"] or 0
                row_end   = add_hours(current_start, msg_hours) if current_start is not None else None
                output.append({
                    "date":        er["date"],
                    "worker_name": er["worker_name"],
                    "workplace":   mr["workplace"],
                    "start_time":  current_start,
                    "end_time":    row_end,
                    "sales":       mr["sales"],
                    "notes":       base_note,
                    "status":      status,
                })
                current_start = row_end
            continue
        # ────────────────────────────────────────────────────────────────────

        count   = max(len(ex_rows), len(ms_rows))

        for i in range(count):
            er = ex_rows[i] if i < len(ex_rows) else None
            mr = ms_rows[i] if i < len(ms_rows) else None

            if er and mr:
                eh, mh = er["hours"], mr["hours"]
                start_t = er["start_time"]
                end_t   = er["end_time"]
                partial_note = None

                # One Excel time missing — calculate it from message hours
                if start_t is None and end_t is not None:
                    partial_note = "שעת התחלה חסרה באקסל"
                    if mh is not None:
                        start_t = subtract_hours(end_t, mh)
                    eh = mh
                elif end_t is None and start_t is not None:
                    partial_note = "שעת סיום חסרה באקסל"
                    if mh is not None:
                        end_t = add_hours(start_t, mh)
                    eh = mh

                if mh is None:
                    # Message line had no hours — use whatever Excel times we have
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  start_t,
                        "end_time":    end_t,
                        "sales":       mr["sales"],
                        "notes":       partial_note or "שעות חסרות בהודעה",
                        "status":      "gap",
                    })
                    continue

                if partial_note:
                    # Times completed from message hours — flag for review
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  start_t,
                        "end_time":    end_t,
                        "sales":       mr["sales"],
                        "notes":       partial_note,
                        "status":      "gap",
                    })
                    continue

                diff_min = abs(eh - mh) * 60

                if diff_min <= threshold_min:
                    # Rule 1 — times match
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  start_t,
                        "end_time":    end_t,
                        "sales":       mr["sales"],
                        "notes":       "הכל תקין",
                        "status":      "ok",
                    })
                else:
                    # Rule 2 — gap: use Excel start, message hours for end
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  start_t,
                        "end_time":    add_hours(start_t, mh),
                        "sales":       mr["sales"],
                        "notes":       f"פער של {int(diff_min)} דקות",
                        "status":      "gap",
                    })

            elif er:
                # Managers are optional — if they're not in the message, skip them silently
                if _is_ignored(er["worker_name"], managers) or _is_ignored(key, managers):
                    continue
                # Rule 3 — missing from message
                output.append({
                    "date":        er["date"],
                    "worker_name": er["worker_name"],
                    "workplace":   "",
                    "start_time":  er["start_time"],
                    "end_time":    er["end_time"],
                    "sales":       "",
                    "notes":       "חסר בהודעה",
                    "status":      "missing_msg",
                })

            else:
                # Rule 4 — missing from Excel
                mh = mr["hours"] or 0
                output.append({
                    "date":        excel_date or "",
                    "worker_name": mr["worker_name"],
                    "workplace":   mr["workplace"],
                    "start_time":  default_start,
                    "end_time":    add_hours(default_start, mh),
                    "sales":       mr["sales"],
                    "notes":       "חסר באקסל",
                    "status":      "missing_excel",
                })

    return output


# ── Export ─────────────────────────────────────────────────────────────────────

def _fmt_time(t) -> str:
    if isinstance(t, time_type):
        return t.strftime("%H:%M")
    return str(t) if t else ""


def _fmt_date(d) -> str:
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y")
    if hasattr(d, "strftime"):
        return d.strftime("%d/%m/%Y")
    return str(d) if (d and not _is_na(d)) else ""


STATUS_COLORS = {
    "ok":            "E2EFDA",  # green
    "gap":           "FFF2CC",  # yellow
    "missing_msg":   "FCE4D6",  # orange
    "missing_excel": "FFDADA",  # red
}

HEADERS = ["תאריך", "שם עובד", "מקום עבודה", "שעת התחלה", "שעת סיום", "מכירות", "הערות"]


def export_to_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "השוואה"
    ws.sheet_view.rightToLeft = True

    ws.append(HEADERS)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            _fmt_date(row["date"]),
            row["worker_name"],
            row["workplace"],
            _fmt_time(row["start_time"]),
            _fmt_time(row["end_time"]),
            row["sales"],
            row["notes"],
        ])
        color = STATUS_COLORS.get(row["status"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 14)

    wb.save(output_path)
